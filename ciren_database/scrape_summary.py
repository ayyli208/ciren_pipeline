"""
Scrapes the crash summaries for each ciren case, and puts it in OUTPUT_FILE
Used later to determine which crash scenario each case falls under
"""
import os
import re
import time

import pyperclip
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

import openpyxl
from openpyxl import load_workbook
import os


DOWNLOAD_FOLDER = rf"D:\UMich\Senior Year\umtri\clean\ciren_database\CrashExports"
OUTPUT_FILE = rf"D:\UMich\Senior Year\umtri\clean\ciren_database\ciren_crash_summaries.xlsx"
PAGE_LOAD_TIMEOUT_SECONDS = 45
MAX_NAV_RETRIES = 2


def build_driver() -> webdriver.Chrome:
    options = Options()
    options.add_argument("--log-level=3")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    if DOWNLOAD_FOLDER:
        os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
        options.add_experimental_option(
            "prefs",
            {
                "download.default_directory": os.path.abspath(DOWNLOAD_FOLDER),
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
            },
        )
    d = webdriver.Chrome(options=options)
    d.set_page_load_timeout(PAGE_LOAD_TIMEOUT_SECONDS)
    return d


def extract_crash_summary_from_clipboard(raw_text: str) -> str:
    if not raw_text:
        return ""
    text = raw_text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)

    # Prefer the Crash Summary block immediately before Injury Analysis.
    injury_match = re.search(r"\bInjury Analysis\b", text, flags=re.IGNORECASE)
    prefix = text[:injury_match.start()] if injury_match else text

    crash_matches = list(re.finditer(r"\bCrash Summary\b", prefix, flags=re.IGNORECASE))
    if not crash_matches:
        return ""

    # Use the last "Crash Summary" before Injury Analysis to avoid nav/header hits.
    start_idx = crash_matches[-1].end()
    summary = prefix[start_idx:]

    # If the first line is just "Summary", drop it.
    summary = re.sub(r"^\s*Summary\s*\n", "", summary, flags=re.IGNORECASE)
    summary = summary.strip()

    # Keep paragraph breaks, but normalize excess blank lines.
    summary = re.sub(r"\n{3,}", "\n\n", summary)
    return summary


def copy_page_text(driver: webdriver.Chrome) -> str:
    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    body = driver.find_element(By.TAG_NAME, "body")
    body.click()
    pyperclip.copy("")
    ActionChains(driver).key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
    time.sleep(0.2)
    ActionChains(driver).key_down(Keys.CONTROL).send_keys("c").key_up(Keys.CONTROL).perform()
    time.sleep(0.4)
    return pyperclip.paste()


def navigate_with_retries(driver: webdriver.Chrome, url: str, retries: int = MAX_NAV_RETRIES) -> bool:
    for attempt in range(1, retries + 1):
        try:
            driver.get(url)
            return True
        except (TimeoutException, WebDriverException) as exc:
            print(f"  Navigation attempt {attempt}/{retries} failed: {exc}")
            try:
                driver.execute_script("window.stop();")
            except Exception:
                pass
            time.sleep(1.0)
    return False


def main() -> None:
    cirenids1 = [11, 15, 16, 17, 19, 20, 28, 30, 31, 34, 35, 39, 41, 42, 43, 45, 49, 51, 56, 57, 58, 63, 64, 66, 67, 68, 69, 73, 76, 77, 78, 80, 84, 89, 90, 91, 92, 95, 98, 99, 100, 102, 103, 104, 105, 110, 112, 115, 117, 119, 123, 132, 133, 134, 135, 136, 138, 139, 141, 145, 150, 158, 160, 161, 162, 164, 180, 181, 189, 193, 194, 195, 197, 198, 200, 201, 211, 214, 216, 220, 221, 226, 227, 229, 230, 231, 244, 248, 251, 253, 258, 262, 266, 267, 274, 287, 290, 291, 298, 299]
    cirenids2 = [303, 310, 311, 324, 341, 350, 351, 352, 359, 363, 398, 406, 408, 409, 417, 420, 421, 424, 426, 427, 428, 432, 433, 434, 439, 440, 444, 459, 460, 465, 479, 490, 497, 518, 527, 533, 536, 537, 542, 550, 555, 557, 558, 559, 567, 580, 581, 584, 590, 594, 597, 623, 634, 653, 658, 661, 664, 665, 666, 678, 681, 687, 702, 704, 708, 709, 718, 725, 730, 731, 732, 733, 740, 742, 743, 748, 759, 760, 761, 769, 783, 798, 800, 802, 804, 805, 806, 811, 814, 816, 818, 819, 824, 826, 827, 828, 853, 854, 866, 868]
    cirenids3 = [871, 882, 883, 897, 898, 915, 916, 923, 938, 945, 948, 962, 963, 972, 977, 980, 982, 984, 990, 1010, 1034, 1047, 1056, 1070, 1078, 1088, 1089, 1094, 1098, 1101, 1124, 1125, 1148, 1149, 1157, 1216, 1217, 1245, 980015, 980018, 980019, 980020, 980024, 980035, 980038, 980039, 980052, 980054, 980090, 980091, 980103, 980112, 980113, 980114, 980119, 980121, 980122, 980145, 980146, 980188, 980217, 980340]
    cirenids = cirenids1 + cirenids2 + cirenids3

    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["cirenid", "crash_summary"])


    driver = build_driver()
    try:
        for case_id in cirenids:
            print(f"Processing CIRENID {case_id}...")
            case_url = f"https://crashviewer.nhtsa.dot.gov/ciren/details/{case_id}/ciren-summary-document"

            ok = navigate_with_retries(driver, case_url)
            if not ok:
                ws.append([case_id, ""])
                wb.save(OUTPUT_FILE)
                print("  Skipped after navigation retries")
                try:
                    driver.quit()
                except Exception:
                    pass
                driver = build_driver()
                continue

            try:
                time.sleep(1.0)
                raw_text = copy_page_text(driver)
                crash_summary = extract_crash_summary_from_clipboard(raw_text)
                ws.append([case_id, crash_summary])
                if crash_summary:
                    print(f"  Extracted summary ({len(crash_summary)} chars)")
                else:
                    print("  Crash summary not found")
            except Exception as exc:
                print(f"  Error extracting summary: {exc}")
                ws.append([case_id, ""])

            wb.save(OUTPUT_FILE)
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    print(f"Done. Results saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
